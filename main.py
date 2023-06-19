from getpass import getpass
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains, Keys
from datetime import datetime as time
from openpyxl import load_workbook
from time import sleep
import itertools

class Sistema:
    def __init__(self, usuario = input('usuario: '), senha = getpass('senha: '), estado_janela=''):
        self.usuario = usuario
        self.senha = senha
        self.service = Service(executable_path="./chromedriver.exe")
        self.navegador = webdriver.Chrome(service=self.service)
        if (estado_janela == 'min'):
            self.navegador.minimize_window()
    
    def login(self):
        if(self.usuario == '' or self.senha == '' ):
            print('usuario e senha nao definidos')
        else:
            self.navegador.get('https://www.sistemas.pa.gov.br/governodigital/public/main/index.xhtml')
            self.navegador.find_element(By.ID, "form_login:login_username").send_keys(self.usuario)
            self.navegador.find_element(By.ID, "form_login:login_password").send_keys(self.senha)
            self.navegador.find_element(By.ID, "form_login:button_login").click()
            WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="form_sistema:submit_area"]/div/div[3]/div[1]/a/img')))
            self.navegador.get('https://www.sistemas.pa.gov.br/sispat')
        
    def acessar_dist_nao_recebido(self):
        dist_nao_recebido = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="form_pendencias:list:1:pendencia_17"]')))
        dist_nao_recebido.click()
    
    def receber(self):
        self.acessar_dist_nao_recebido()    
        while True:
            transferencia_nao_recebida_btn = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/table/tbody/tr/td[3]/div/form[2]/span/table/tbody/tr[1]/td[7]')))
            transferencia_nao_recebida_btn.click()
            calendario_btn = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/table/tbody/tr[2]/td/form/fieldset/div/table/tbody/tr[4]/td[2]/span/img')))
            calendario_btn.click()
            data_btn = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/table/tbody/tr[2]/td/form/fieldset/div/table/tbody/tr[4]/td[2]/table/tbody/tr[4]/td[4]')))
            data_btn.click()
            receber = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/table/tbody/tr[2]/td/form/div/input[1]')))
            receber.click()
            WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/table/tbody/tr/td[3]/div/form[2]/span/table/tbody/tr[1]/td[7]')))
        
    def acessar_entrada_por_transferência_nao_incorporado(self):
        entrada_por_transferência_nao_incorporado = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/table/tbody/tr/td[3]/div/form[3]/div/table/tbody/tr/td[1]/a')))
        entrada_por_transferência_nao_incorporado.click()
        

        
    def incorporar(self, descricao=''):
        self.acessar_entrada_por_transferência_nao_incorporado()
        if(descricao != ''):
            descricao_input = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="incorporar_bem_destinado_ao_orgao_form_pesq:descricaomaterial"]')))
            descricao_input.send_keys(descricao)
            pesquisar = self.navegador.find_element(By.XPATH, '//*[@id="incorporar_bem_destinado_ao_orgao_form_pesq:j_id433"]')
            pesquisar.click()
            sleep(4)
            
            
            
        cadastrados = 0
        wb_patrimonio = load_workbook(filename = 'patrimonios.xlsx')
        wb_incorporados = load_workbook(filename = 'incorporados.xlsx')
        patrimonio = wb_patrimonio.active
        incorporados = wb_incorporados.active
        
        total = 0
        for row in patrimonio.values:   
            if(row[0] != None):
                total += 1  
        
        cells = incorporados.values

        cells_preenchidas = 0
        for row in cells:
            if(row[0] != None):
                cells_preenchidas += 1
        
        if(cells_preenchidas == 0):
            cells_preenchidas = 1
        
        rps_para_incorporar = patrimonio[1:total]
        rps_incorporados = incorporados[cells_preenchidas:cells_preenchidas+total]
        
        for row_rp, row_incorporado in zip(rps_para_incorporar, rps_incorporados):
            rp = row_rp.value
            if (rp != None):
                row_incorporado[0].value = rp
                cadastrados += 1
                
                cadastrar = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/table/tbody/tr/td[3]/div/form[2]/span/table/tbody/tr[1]/td[8]/a/img')))
                cadastrar.click()
                
                input_rp = WebDriverWait(self.navegador, timeout=30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/table/tbody/tr/td[3]/div/form/div/div/table[2]/tbody/tr[2]/td[2]/input')))
                input_rp.send_keys("")
                ActionChains(self.navegador).send_keys(Keys.TAB).send_keys(Keys.ENTER).perform()
                confirmacao = WebDriverWait(self.navegador, timeout=60).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/table/tbody/tr/td[3]/div/div[1]/table/tbody/tr/td/span[2]')))
                assert confirmacao.text == "Bem foi incorporado ao órgão com sucesso."
                
                timestamp = time.now()
                print(f"\33[1;96m{timestamp}\33[1;37m - Patrimonio: \33[1;35m{rp}\33[92m OK\33[1;37m - Progresso: \33[1;96m{cadastrados}/{total}\33[m\n")
                row_rp[0].value = ''
                
                wb_incorporados.save('incorporados.xlsx')
                wb_patrimonio.save('patrimonios.xlsx')
                    

if __name__ == '__main__':
    sispat = Sistema('yago.martins', '7366yawi')
    sispat.login()
    sispat.incorporar()
    